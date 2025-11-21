"""Tests for the storage abstraction layer."""

import pytest
import json
import csv
from pathlib import Path

from src.storage import StorageFactory, StorageBackend


class TestStorageFactory:
    """Tests for StorageFactory."""

    def test_available_formats(self):
        """Test listing available storage formats."""
        formats = StorageFactory.available_formats()
        assert "json" in formats
        assert "csv" in formats
        assert "xlsx" in formats
        assert "excel" in formats

    def test_create_json_storage(self, temp_dir):
        """Test creating JSON storage backend."""
        storage = StorageFactory.create("json", temp_dir)
        assert storage is not None
        assert isinstance(storage, StorageBackend)

    def test_create_csv_storage(self, temp_dir):
        """Test creating CSV storage backend."""
        storage = StorageFactory.create("csv", temp_dir)
        assert storage is not None
        assert isinstance(storage, StorageBackend)

    def test_create_excel_storage(self, temp_dir):
        """Test creating Excel storage backend."""
        storage = StorageFactory.create("xlsx", temp_dir)
        assert storage is not None
        assert isinstance(storage, StorageBackend)

    def test_create_unknown_format_raises_error(self, temp_dir):
        """Test that unknown format raises ValueError."""
        with pytest.raises(ValueError) as exc_info:
            StorageFactory.create("unknown", temp_dir)
        assert "Unknown storage format" in str(exc_info.value)


class TestJSONStorage:
    """Tests for JSON storage backend."""

    @pytest.mark.unit
    def test_save_posts(self, json_storage, sample_posts):
        """Test saving posts to JSON."""
        filepath = json_storage.save_posts(sample_posts, "test_account", "instagram")

        assert Path(filepath).exists()
        assert filepath.endswith(".json")

        with open(filepath) as f:
            data = json.load(f)

        assert len(data) == len(sample_posts)
        assert data[0]["platform_id"] == sample_posts[0].platform_id
        assert data[0]["account_id"] == sample_posts[0].account_id

    @pytest.mark.unit
    def test_save_comments(self, json_storage, sample_comments):
        """Test saving comments to JSON."""
        filepath = json_storage.save_comments(sample_comments, "test_account", "instagram")

        assert Path(filepath).exists()

        with open(filepath) as f:
            data = json.load(f)

        assert len(data) == len(sample_comments)
        assert data[0]["platform_id"] == sample_comments[0].platform_id

    @pytest.mark.unit
    def test_save_profile(self, json_storage, sample_profile):
        """Test saving profile to JSON."""
        filepath = json_storage.save_profile(sample_profile, "test_account", "instagram")

        assert Path(filepath).exists()

        with open(filepath) as f:
            data = json.load(f)

        assert data["username"] == sample_profile.username
        assert data["followers_count"] == sample_profile.followers_count

    @pytest.mark.unit
    def test_save_empty_posts(self, json_storage):
        """Test saving empty posts list."""
        filepath = json_storage.save_posts([], "test_account", "instagram")

        assert Path(filepath).exists()

        with open(filepath) as f:
            data = json.load(f)

        assert data == []


class TestCSVStorage:
    """Tests for CSV storage backend."""

    @pytest.mark.unit
    def test_save_posts(self, csv_storage, sample_posts):
        """Test saving posts to CSV."""
        filepath = csv_storage.save_posts(sample_posts, "test_account", "instagram")

        assert Path(filepath).exists()
        assert filepath.endswith(".csv")

        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == len(sample_posts)
        assert rows[0]["platform_id"] == sample_posts[0].platform_id

    @pytest.mark.unit
    def test_save_comments(self, csv_storage, sample_comments):
        """Test saving comments to CSV."""
        filepath = csv_storage.save_comments(sample_comments, "test_account", "instagram")

        assert Path(filepath).exists()

        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == len(sample_comments)

    @pytest.mark.unit
    def test_save_profile(self, csv_storage, sample_profile):
        """Test saving profile to CSV."""
        filepath = csv_storage.save_profile(sample_profile, "test_account", "instagram")

        assert Path(filepath).exists()

        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 1
        assert rows[0]["username"] == sample_profile.username


class TestExcelStorage:
    """Tests for Excel storage backend."""

    @pytest.mark.unit
    def test_save_posts(self, excel_storage, sample_posts):
        """Test saving posts to Excel."""
        filepath = excel_storage.save_posts(sample_posts, "test_account", "instagram")

        assert Path(filepath).exists()
        assert filepath.endswith(".xlsx")

    @pytest.mark.unit
    def test_save_comments(self, excel_storage, sample_comments):
        """Test saving comments to Excel."""
        filepath = excel_storage.save_comments(sample_comments, "test_account", "instagram")

        assert Path(filepath).exists()

    @pytest.mark.unit
    def test_save_profile(self, excel_storage, sample_profile):
        """Test saving profile to Excel."""
        filepath = excel_storage.save_profile(sample_profile, "test_account", "instagram")

        assert Path(filepath).exists()

    @pytest.mark.unit
    def test_save_extraction_result(self, excel_storage, sample_posts, sample_comments, sample_profile):
        """Test saving complete extraction result to Excel."""
        result = excel_storage.save_extraction_result(
            posts=sample_posts,
            comments=sample_comments,
            profile=sample_profile,
            account="test_account",
            platform="instagram"
        )

        assert "extraction" in result
        assert Path(result["extraction"]).exists()

        # Verify it's an Excel file with multiple sheets
        from openpyxl import load_workbook
        wb = load_workbook(result["extraction"])

        assert "Posts" in wb.sheetnames
        assert "Comments" in wb.sheetnames
        assert "Profile" in wb.sheetnames


class TestStorageBackendConversions:
    """Tests for data conversion methods in StorageBackend."""

    @pytest.mark.unit
    def test_post_to_dict(self, json_storage, sample_post):
        """Test converting Post to dictionary."""
        data = json_storage.post_to_dict(sample_post)

        assert data["platform_id"] == sample_post.platform_id
        assert data["account_id"] == sample_post.account_id
        assert data["text"] == sample_post.text
        assert data["likes"] == sample_post.likes
        assert data["published_at"] is not None

    @pytest.mark.unit
    def test_comment_to_dict(self, json_storage, sample_comment):
        """Test converting Comment to dictionary."""
        data = json_storage.comment_to_dict(sample_comment)

        assert data["platform_id"] == sample_comment.platform_id
        assert data["post_id"] == sample_comment.post_id
        assert data["author"]["username"] == sample_comment.author.username
        assert data["text"] == sample_comment.text
        assert data["likes"] == sample_comment.likes

    @pytest.mark.unit
    def test_profile_to_dict(self, json_storage, sample_profile):
        """Test converting Profile to dictionary."""
        data = json_storage.profile_to_dict(sample_profile)

        assert data["username"] == sample_profile.username
        assert data["display_name"] == sample_profile.display_name
        assert data["followers_count"] == sample_profile.followers_count
        assert data["is_verified"] == sample_profile.is_verified


class TestStorageExtractionResult:
    """Tests for save_extraction_result method."""

    @pytest.mark.unit
    def test_extraction_result_with_all_data(self, json_storage, sample_posts, sample_comments, sample_profile):
        """Test saving complete extraction result."""
        result = json_storage.save_extraction_result(
            posts=sample_posts,
            comments=sample_comments,
            profile=sample_profile,
            account="test_account",
            platform="instagram"
        )

        assert "posts" in result
        assert "comments" in result
        assert "profile" in result
        assert Path(result["posts"]).exists()
        assert Path(result["comments"]).exists()
        assert Path(result["profile"]).exists()

    @pytest.mark.unit
    def test_extraction_result_with_empty_data(self, json_storage):
        """Test saving extraction result with empty data."""
        result = json_storage.save_extraction_result(
            posts=[],
            comments=[],
            profile=None,
            account="test_account",
            platform="instagram"
        )

        # With empty data, no files should be created
        assert "posts" not in result
        assert "comments" not in result
        assert "profile" not in result
