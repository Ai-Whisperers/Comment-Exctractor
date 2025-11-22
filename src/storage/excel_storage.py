"""Excel storage backend implementation."""

import logging
from typing import List, Dict, Any, Optional, Tuple, Type

from .base import StorageBackend, StorageFactory
from ..core.models import Post, Comment, Profile

logger = logging.getLogger(__name__)


class ExcelStorage(StorageBackend):
    """Excel file storage backend using openpyxl."""

    def _get_workbook(self) -> Tuple[Type, Type, Type, Type]:
        """Import openpyxl lazily to avoid import errors if not installed."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            return Workbook, Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

    def _apply_header_style(self, ws: Any, headers: List[str]) -> None:
        """Apply styling to header row."""
        _, Font, PatternFill, Alignment = self._get_workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _auto_fit_columns(self, ws: Any) -> None:
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(50, max(10, max_length + 2))
            ws.column_dimensions[column_letter].width = adjusted_width

    def save_posts(
        self,
        posts: List[Post],
        account: str,
        platform: str
    ) -> str:
        """Save posts to Excel file."""
        Workbook, _, _, _ = self._get_workbook()
        filepath = self._generate_filename(account, platform, "posts", "xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Posts"

        headers = [
            "Platform ID", "Account ID", "Text", "Likes",
            "Comments", "Shares", "URL", "Media Type",
            "Published At", "Platform"
        ]
        self._apply_header_style(ws, headers)

        for row, post in enumerate(posts, 2):
            data = self.post_to_dict(post)
            ws.cell(row=row, column=1, value=data["platform_id"])
            ws.cell(row=row, column=2, value=data["account_id"])
            ws.cell(row=row, column=3, value=data["text"])
            ws.cell(row=row, column=4, value=data["likes"])
            ws.cell(row=row, column=5, value=data["comments_count"])
            ws.cell(row=row, column=6, value=data["shares"])
            ws.cell(row=row, column=7, value=data["url"])
            ws.cell(row=row, column=8, value=data["media_type"])
            ws.cell(row=row, column=9, value=data["published_at"])
            ws.cell(row=row, column=10, value=data["platform"])

        self._auto_fit_columns(ws)
        wb.save(filepath)

        logger.debug(f"Saved {len(posts)} posts to {filepath}")
        return str(filepath)

    def save_comments(
        self,
        comments: List[Comment],
        account: str,
        platform: str
    ) -> str:
        """Save comments to Excel file."""
        Workbook, _, _, _ = self._get_workbook()
        filepath = self._generate_filename(account, platform, "comments", "xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Comments"

        headers = [
            "Platform ID", "Post ID", "Author", "Text", "Likes",
            "Parent ID", "Replies", "Published At", "Platform"
        ]
        self._apply_header_style(ws, headers)

        for row, comment in enumerate(comments, 2):
            data = self.comment_to_dict(comment)
            author_str = data["author"]["username"] if data["author"] else ""
            ws.cell(row=row, column=1, value=data["platform_id"])
            ws.cell(row=row, column=2, value=data["post_id"])
            ws.cell(row=row, column=3, value=author_str)
            ws.cell(row=row, column=4, value=data["text"])
            ws.cell(row=row, column=5, value=data["likes"])
            ws.cell(row=row, column=6, value=data["parent_id"])
            ws.cell(row=row, column=7, value=data["replies_count"])
            ws.cell(row=row, column=8, value=data["published_at"])
            ws.cell(row=row, column=9, value=data["platform"])

        self._auto_fit_columns(ws)
        wb.save(filepath)

        logger.debug(f"Saved {len(comments)} comments to {filepath}")
        return str(filepath)

    def save_profile(
        self,
        profile: Profile,
        account: str,
        platform: str
    ) -> str:
        """Save profile to Excel file."""
        Workbook, _, _, _ = self._get_workbook()
        filepath = self._generate_filename(account, platform, "profile", "xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Profile"

        headers = [
            "Platform ID", "Username", "Display Name", "Description", "Followers",
            "Following", "Posts", "URL", "Profile Image", "Verified", "Platform"
        ]
        self._apply_header_style(ws, headers)

        data = self.profile_to_dict(profile)
        ws.cell(row=2, column=1, value=data["platform_id"])
        ws.cell(row=2, column=2, value=data["username"])
        ws.cell(row=2, column=3, value=data["display_name"])
        ws.cell(row=2, column=4, value=data["description"])
        ws.cell(row=2, column=5, value=data["followers_count"])
        ws.cell(row=2, column=6, value=data["following_count"])
        ws.cell(row=2, column=7, value=data["posts_count"])
        ws.cell(row=2, column=8, value=data["url"])
        ws.cell(row=2, column=9, value=data["profile_image"])
        ws.cell(row=2, column=10, value=data["is_verified"])
        ws.cell(row=2, column=11, value=data["platform"])

        self._auto_fit_columns(ws)
        wb.save(filepath)

        logger.debug(f"Saved profile to {filepath}")
        return str(filepath)

    def save_extraction_result(
        self,
        posts: List[Post],
        comments: List[Comment],
        profile: Optional[Profile],
        account: str,
        platform: str
    ) -> Dict[str, str]:
        """Save complete extraction to a single Excel file with multiple sheets."""
        Workbook, _, _, _ = self._get_workbook()
        filepath = self._generate_filename(account, platform, "extraction", "xlsx")

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active

        # Posts sheet
        if posts:
            ws_posts = wb.create_sheet("Posts")
            headers = [
                "Platform ID", "Account ID", "Text", "Likes",
                "Comments", "Shares", "URL", "Media Type", "Published At", "Platform"
            ]
            self._apply_header_style(ws_posts, headers)

            for row, post in enumerate(posts, 2):
                data = self.post_to_dict(post)
                ws_posts.cell(row=row, column=1, value=data["platform_id"])
                ws_posts.cell(row=row, column=2, value=data["account_id"])
                ws_posts.cell(row=row, column=3, value=data["text"])
                ws_posts.cell(row=row, column=4, value=data["likes"])
                ws_posts.cell(row=row, column=5, value=data["comments_count"])
                ws_posts.cell(row=row, column=6, value=data["shares"])
                ws_posts.cell(row=row, column=7, value=data["url"])
                ws_posts.cell(row=row, column=8, value=data["media_type"])
                ws_posts.cell(row=row, column=9, value=data["published_at"])
                ws_posts.cell(row=row, column=10, value=data["platform"])

            self._auto_fit_columns(ws_posts)

        # Comments sheet
        if comments:
            ws_comments = wb.create_sheet("Comments")
            headers = [
                "Platform ID", "Post ID", "Author", "Text", "Likes",
                "Parent ID", "Replies", "Published At", "Platform"
            ]
            self._apply_header_style(ws_comments, headers)

            for row, comment in enumerate(comments, 2):
                data = self.comment_to_dict(comment)
                author_str = data["author"]["username"] if data["author"] else ""
                ws_comments.cell(row=row, column=1, value=data["platform_id"])
                ws_comments.cell(row=row, column=2, value=data["post_id"])
                ws_comments.cell(row=row, column=3, value=author_str)
                ws_comments.cell(row=row, column=4, value=data["text"])
                ws_comments.cell(row=row, column=5, value=data["likes"])
                ws_comments.cell(row=row, column=6, value=data["parent_id"])
                ws_comments.cell(row=row, column=7, value=data["replies_count"])
                ws_comments.cell(row=row, column=8, value=data["published_at"])
                ws_comments.cell(row=row, column=9, value=data["platform"])

            self._auto_fit_columns(ws_comments)

        # Profile sheet
        if profile:
            ws_profile = wb.create_sheet("Profile")
            headers = [
                "Platform ID", "Username", "Display Name", "Description", "Followers",
                "Following", "Posts", "URL", "Profile Image", "Verified", "Platform"
            ]
            self._apply_header_style(ws_profile, headers)

            data = self.profile_to_dict(profile)
            ws_profile.cell(row=2, column=1, value=data["platform_id"])
            ws_profile.cell(row=2, column=2, value=data["username"])
            ws_profile.cell(row=2, column=3, value=data["display_name"])
            ws_profile.cell(row=2, column=4, value=data["description"])
            ws_profile.cell(row=2, column=5, value=data["followers_count"])
            ws_profile.cell(row=2, column=6, value=data["following_count"])
            ws_profile.cell(row=2, column=7, value=data["posts_count"])
            ws_profile.cell(row=2, column=8, value=data["url"])
            ws_profile.cell(row=2, column=9, value=data["profile_image"])
            ws_profile.cell(row=2, column=10, value=data["is_verified"])
            ws_profile.cell(row=2, column=11, value=data["platform"])

            self._auto_fit_columns(ws_profile)

        # Remove default sheet if we have other sheets
        if len(wb.sheetnames) > 1:
            wb.remove(default_sheet)

        wb.save(filepath)

        logger.info(f"Saved complete extraction to {filepath}")
        return {"extraction": str(filepath)}


# Register the backend
StorageFactory.register("xlsx", ExcelStorage)
StorageFactory.register("excel", ExcelStorage)
